"""Export records/accounts.json + records/model-history.jsonl into one
modern, categorised, READ-ONLY Excel workbook: records/exports/records.xlsx.

Sheets: Users | Rewards | Invoices | Predictions | Results | Model Performance.
Every sheet is protected (no editing), has frozen headers and auto-filters.
Run:  .venv/bin/python model/src/export_records.py
"""
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
EXPORT_DIR = ROOT / "records" / "exports"

NAVY = "1B2A4A"
ACCENT = "2ECC71"
STRIPE = "F4F6FA"
PROTECT_PW = "vendicator-readonly"


def _style_sheet(ws, headers, rows, tab_color=NAVY):
    ws.sheet_properties.tabColor = tab_color
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for r, row in enumerate(rows, start=2):
        ws.append(row)
        if r % 2 == 0:
            for cell in ws[r]:
                cell.fill = PatternFill("solid", fgColor=STRIPE)
    for i, h in enumerate(headers, start=1):
        width = max([len(str(h))] +
                    [len(str(row[i - 1])) for row in rows if len(row) >= i]
                    or [10])
        ws.column_dimensions[get_column_letter(i)].width = min(width + 4, 40)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    ws.protection.sheet = True          # read-only
    ws.protection.password = PROTECT_PW
    ws.protection.autoFilter = False    # keep filter dropdowns usable
    ws.protection.sort = False


def load_records():
    accounts = json.loads((ROOT / "records" / "accounts.json").read_text())
    history = []
    hist_path = ROOT / "records" / "model-history.jsonl"
    if hist_path.exists():
        for line in hist_path.read_text().splitlines():
            if line.strip():
                history.append(json.loads(line))
    return accounts, history


def build(accounts, history):
    wb = Workbook()

    users = [u for u in accounts.get("users", [])]
    ws = wb.active
    ws.title = "Users"
    _style_sheet(ws, ["ID", "Username", "Email", "Joined", "Tier", "Points",
                      "Lifetime Points", "Rank", "Top 5 Players", "Top 5 Teams"],
                 [[u.get("id"), u.get("username"), u.get("email"),
                   u.get("joined"), u.get("tier"), u.get("points_balance"),
                   u.get("lifetime_points"), u.get("ranking"),
                   ", ".join(u.get("top5_players", [])),
                   ", ".join(u.get("top5_teams", []))] for u in users])

    _style_sheet(wb.create_sheet("Rewards"),
                 ["User", "Reward", "Points Spent", "Date", "Status"],
                 [[u.get("username"), r.get("reward"), r.get("points"),
                   r.get("date"), r.get("status")]
                  for u in users for r in u.get("rewards_redeemed", [])],
                 tab_color=ACCENT)

    subs = accounts.get("subscriptions", [])
    _style_sheet(wb.create_sheet("Subscriptions"),
                 ["Subscription ID", "User", "Tier", "Started",
                  "Unlocked By Points", "Status", "Benefits"],
                 [[s.get("subscription_id"), s.get("user_id"),
                   s.get("tier"), s.get("started"),
                   s.get("unlocked_by_points"), s.get("status"),
                   "; ".join(s.get("benefits_snapshot", []))]
                  for s in subs], tab_color=ACCENT)

    _style_sheet(wb.create_sheet("Invoices"),
                 ["Invoice ID", "User", "Date", "Amount", "Currency",
                  "Description", "Status"],
                 [[i.get("invoice_id"), u.get("username"), i.get("date"),
                   i.get("amount"), i.get("currency"), i.get("description"),
                   i.get("status")]
                  for u in users for i in u.get("invoices", [])])

    preds = [h for h in history if not h.get("settled")]
    _style_sheet(wb.create_sheet("Predictions"),
                 ["Timestamp", "League", "Match", "Kickoff", "Model", "Market",
                  "Prediction", "Probability %", "Fair Odds", "Book Odds"],
                 [[h.get("ts"), h.get("league"),
                   f"{h.get('home')} vs {h.get('away')}", h.get("kickoff"),
                   h.get("model"), h.get("market"),
                   json.dumps(h.get("prediction")),
                   json.dumps(h.get("probs")), json.dumps(h.get("fair_odds")),
                   json.dumps(h.get("book_odds"))] for h in preds])

    settled = [h for h in history if h.get("settled")]
    _style_sheet(wb.create_sheet("Results"),
                 ["Timestamp", "League", "Match", "Model", "Market",
                  "Prediction", "Result", "Brier", "Log Loss", "CLV"],
                 [[h.get("ts"), h.get("league"),
                   f"{h.get('home')} vs {h.get('away')}", h.get("model"),
                   h.get("market"), json.dumps(h.get("prediction")),
                   json.dumps(h.get("result")),
                   (h.get("eval") or {}).get("brier"),
                   (h.get("eval") or {}).get("logloss"),
                   (h.get("eval") or {}).get("clv")] for h in settled])

    perf_rows = []
    by_model = {}
    for h in settled:
        ev = h.get("eval") or {}
        if ev.get("brier") is not None:
            by_model.setdefault(h.get("model"), []).append(ev)
    for model, evs in sorted(by_model.items()):
        briers = [e["brier"] for e in evs if e.get("brier") is not None]
        lls = [e["logloss"] for e in evs if e.get("logloss") is not None]
        clvs = [e["clv"] for e in evs if e.get("clv") is not None]
        perf_rows.append([
            model, len(evs),
            round(sum(briers) / len(briers), 4) if briers else None,
            round(sum(lls) / len(lls), 4) if lls else None,
            round(sum(clvs) / len(clvs), 4) if clvs else None])
    _style_sheet(wb.create_sheet("Model Performance"),
                 ["Model", "Settled Predictions", "Avg Brier", "Avg Log Loss",
                  "Avg CLV"], perf_rows, tab_color=ACCENT)

    return wb


def main():
    accounts, history = load_records()
    wb = build(accounts, history)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORT_DIR / "records.xlsx"
    wb.save(out)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    print(f"Wrote {out} ({stamp}) - sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
